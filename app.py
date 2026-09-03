import os
import io
import zipfile
import mimetypes
import base64
import uuid
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, send_from_directory, abort, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
import database
import notifier

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "aquashield_helpdesk_secret_key_2026_aquachile_secure")
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 60 * 60 * 24 * 90 # 90 dias de persistencia

CORS(app, supports_credentials=True)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB max total upload

database.init_db()

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/admin")
def admin():
    return render_template("admin.html")

# ── API: Autenticación & Perfil de Usuario ──────────────────────────────────

@app.route("/api/auth/register", methods=["POST"])
def auth_register():
    try:
        data = request.get_json() or {}
        name = data.get("name", "").strip()
        email = data.get("email", "").strip()
        password = data.get("password", "").strip()
        phone = data.get("phone", "").strip()
        department = data.get("department", "").strip()

        if not name or not email or not password:
            return jsonify({"success": False, "error": "Nombre, correo y contraseña son obligatorios"}), 400

        user = database.register_user(name, email, password, phone, department)
        session.permanent = True
        session["user_id"] = user["id"]
        return jsonify({"success": True, "user": user}), 201
    except ValueError as ve:
        return jsonify({"success": False, "error": str(ve)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    try:
        data = request.get_json() or {}
        email = data.get("email", "").strip()
        password = data.get("password", "").strip()

        if not email or not password:
            return jsonify({"success": False, "error": "Ingresa tu correo y contraseña"}), 400

        user = database.authenticate_user(email, password)
        if not user:
            return jsonify({"success": False, "error": "Credenciales inválidas. Revisa tu correo o contraseña."}), 401

        session.permanent = True
        session["user_id"] = user["id"]
        return jsonify({"success": True, "user": user})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/auth/me", methods=["GET"])
def auth_me():
    user_id = session.get("user_id")
    if not user_id:
        # También permitir buscar por email en query param si se usa almacenamiento local
        req_email = request.args.get("email")
        if req_email:
            u = database.get_user_by_email(req_email)
            if u:
                return jsonify({"authenticated": True, "user": u})
        return jsonify({"authenticated": False, "user": None})

    user = database.get_user_by_id(user_id)
    if not user:
        session.pop("user_id", None)
        return jsonify({"authenticated": False, "user": None})

    return jsonify({"authenticated": True, "user": user})

@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"success": True})

# ── API: Gestión de Usuarios (Admin) ───────────────────────────────────────

@app.route("/api/admin/users", methods=["GET"])
def admin_get_users():
    try:
        search = request.args.get("search", "")
        users = database.get_all_users(search=search)
        return jsonify({"success": True, "users": users})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/admin/users", methods=["POST"])
def admin_create_user():
    try:
        data = request.get_json() or {}
        name = data.get("name", "").strip()
        email = data.get("email", "").strip()
        password = data.get("password", "").strip()
        phone = data.get("phone", "").strip()
        department = data.get("department", "").strip()
        role = data.get("role", "usuario")
        
        if not name or not email or not password:
            return jsonify({"success": False, "error": "Nombre, correo y contraseña son obligatorios"}), 400
            
        user = database.register_user(name, email, password, phone, department, role)
        return jsonify({"success": True, "user": user}), 201
    except ValueError as ve:
        return jsonify({"success": False, "error": str(ve)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/admin/users/<int:user_id>", methods=["PATCH"])
def admin_update_user_endpoint(user_id):
    try:
        data = request.get_json() or {}
        user = database.admin_update_user(user_id, data)
        return jsonify({"success": True, "user": user})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/admin/users/<int:user_id>/reset-password", methods=["POST"])
def admin_reset_password_endpoint(user_id):
    try:
        data = request.get_json() or {}
        new_password = data.get("new_password", "").strip()
        if not new_password:
            return jsonify({"success": False, "error": "Debes especificar la nueva contraseña"}), 400
            
        user = database.admin_reset_user_password(user_id, new_password)
        return jsonify({"success": True, "message": f"Contraseña de {user['name']} actualizada exitosamente."})
    except ValueError as ve:
        return jsonify({"success": False, "error": str(ve)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/admin/users/<int:user_id>", methods=["DELETE"])
def admin_delete_user_endpoint(user_id):
    try:
        res = database.admin_delete_user(user_id)
        return jsonify(res)
    except ValueError as ve:
        return jsonify({"success": False, "error": str(ve)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Configuraciones del Sistema y Correo (Admin) ──────────────────────

@app.route("/api/admin/settings", methods=["GET"])
def admin_get_settings():
    try:
        settings = database.get_all_settings()
        return jsonify({"success": True, "settings": settings})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/admin/settings", methods=["POST"])
def admin_save_settings():
    try:
        data = request.get_json() or {}
        updated = database.update_settings(data)
        return jsonify({"success": True, "settings": updated, "message": "Configuración guardada exitosamente."})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/admin/settings/test-email", methods=["POST"])
def admin_test_email():
    try:
        data = request.get_json() or {}
        if data.get("settings"):
            database.update_settings(data.get("settings", {}))
            
        target_email = data.get("target_email")
        success, msg = notifier.test_email_connection(target_email)
        return jsonify({"success": success, "message": msg})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Módulos ────────────────────────────────────────────────────────────

@app.route("/api/modules", methods=["GET"])
def get_modules_endpoint():
    try:
        active_only = request.args.get("all", "0") != "1"
        modules = database.get_modules(active_only=active_only)
        return jsonify({"success": True, "modules": modules})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/modules", methods=["POST"])
def create_module_endpoint():
    try:
        data = request.get_json() or {}
        name = data.get("name", "").strip()
        description = data.get("description", "").strip()
        
        if not name:
            return jsonify({"success": False, "error": "El nombre del módulo es obligatorio"}), 400
            
        module = database.create_module(name, description)
        return jsonify({"success": True, "module": module}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/modules/<int:module_id>", methods=["DELETE"])
def delete_module_endpoint(module_id):
    try:
        res = database.delete_module(module_id)
        return jsonify(res)
    except ValueError as ve:
        return jsonify({"success": False, "error": str(ve)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Tickets ────────────────────────────────────────────────────────────

@app.route("/api/tickets", methods=["POST"])
def create_ticket_endpoint():
    try:
        if request.is_json:
            data = request.get_json()
            uploaded_files = []
        else:
            data = request.form.to_dict()
            uploaded_files = request.files.getlist("attachments") or request.files.getlist("files[]")

        # Vincular usuario autenticado si existe en sesión
        current_user = None
        user_id = session.get("user_id") or data.get("user_id")
        if user_id:
            current_user = database.get_user_by_id(int(user_id))
            if current_user:
                data["user_id"] = current_user["id"]
                if not data.get("requester_name"):
                    data["requester_name"] = current_user["name"]
                if not data.get("requester_email"):
                    data["requester_email"] = current_user["email"]
                if not data.get("requester_phone") and current_user.get("phone"):
                    data["requester_phone"] = current_user["phone"]

        requester_name = data.get("requester_name", "").strip()
        requester_email = data.get("requester_email", "").strip()
        title = data.get("title", "").strip()
        description = data.get("description", "").strip()

        if not requester_name or not requester_email or not title or not description:
            return jsonify({
                "success": False, 
                "error": "Por favor completa los campos obligatorios (Nombre, Email, Asunto y Descripción)"
            }), 400

        # Si el usuario no existe en la base de datos de usuarios, auto-registrarlo de forma transparente
        if not data.get("user_id"):
            existing_user = database.get_user_by_email(requester_email)
            if existing_user:
                data["user_id"] = existing_user["id"]
            else:
                try:
                    # Crear usuario básico con contraseña temporal por si quiere iniciar sesión
                    new_u = database.register_user(
                        name=requester_name,
                        email=requester_email,
                        password="user" + datetime.now().strftime("%Y%m%d"),
                        phone=data.get("requester_phone", "")
                    )
                    data["user_id"] = new_u["id"]
                except Exception:
                    pass

        # Procesar archivos adjuntos
        temp_files_to_save = []
        for file in uploaded_files:
            if file and file.filename and file.filename.strip():
                original_filename = os.path.basename(file.filename)
                safe_name = secure_filename(original_filename)
                if not safe_name:
                    safe_name = f"adjunto_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                
                mime = file.content_type or mimetypes.guess_type(original_filename)[0] or "application/octet-stream"
                content = file.read()
                file_size = len(content)
                
                temp_files_to_save.append({
                    "original_name": original_filename,
                    "safe_name": safe_name,
                    "mime_type": mime,
                    "file_size": file_size,
                    "content": content
                })

        temp_attachment_records = []
        for f in temp_files_to_save:
            temp_attachment_records.append({
                "original_name": f["original_name"],
                "stored_name": f["safe_name"],
                "file_size": f["file_size"],
                "mime_type": f["mime_type"],
                "file_path": ""
            })

        ticket = database.create_ticket(data, temp_attachment_records)
        ticket_code = ticket["code"]
        
        ticket_upload_dir = os.path.join(UPLOAD_FOLDER, ticket_code)
        os.makedirs(ticket_upload_dir, exist_ok=True)

        with database.get_db() as conn:
            cursor = conn.cursor()
            for idx, f in enumerate(temp_files_to_save):
                stored_filename = f"{idx+1}_{f['safe_name']}"
                disk_path = os.path.join(ticket_upload_dir, stored_filename)
                
                with open(disk_path, "wb") as disk_file:
                    disk_file.write(f["content"])
                    
                rel_path = os.path.relpath(disk_path, BASE_DIR)
                cursor.execute("""
                    UPDATE attachments 
                    SET stored_name = ?, file_path = ? 
                    WHERE ticket_id = ? AND original_name = ?
                """, (stored_filename, rel_path, ticket["id"], f["original_name"]))
            conn.commit()

        final_ticket = database.get_ticket_by_id_or_code(ticket["id"])
        
        # Enviar notificación por correo en segundo plano
        try:
            base_url = request.host_url.rstrip('/')
            notifier.notify_ticket_created(final_ticket, base_url)
        except Exception as e:
            pass

        return jsonify({"success": True, "ticket": final_ticket}), 201

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets", methods=["GET"])
def get_tickets_endpoint():
    try:
        status = request.args.get("status")
        module_id = request.args.get("module_id")
        ticket_type = request.args.get("type")
        priority = request.args.get("priority")
        search = request.args.get("search")
        user_id = request.args.get("user_id")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        tag = request.args.get("tag")
        assigned_to = request.args.get("assigned_to")
        limit = int(request.args.get("limit", 100))
        offset = int(request.args.get("offset", 0))
        
        tickets = database.get_tickets(
            status=status,
            module_id=module_id,
            ticket_type=ticket_type,
            priority=priority,
            search=search,
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            tag=tag,
            assigned_to=assigned_to,
            limit=limit,
            offset=offset
        )
        return jsonify({"success": True, "tickets": tickets})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets/<int:ticket_id>/pin", methods=["POST"])
def toggle_pin_endpoint(ticket_id):
    try:
        res = database.toggle_pin_ticket(ticket_id)
        return jsonify(res)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets/<int:ticket_id>/tags", methods=["POST"])
def update_tags_endpoint(ticket_id):
    try:
        data = request.get_json() or {}
        tags = data.get("tags", "")
        ticket = database.update_ticket_tags(ticket_id, tags)
        return jsonify({"success": True, "ticket": ticket})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/my-tickets", methods=["GET"])
def get_my_tickets_endpoint():
    try:
        user_id = session.get("user_id")
        user_email = request.args.get("email")
        
        if user_id:
            tickets = database.get_tickets(user_id=user_id)
        elif user_email:
            u = database.get_user_by_email(user_email)
            if u:
                tickets = database.get_tickets(user_id=u["id"])
            else:
                tickets = database.get_tickets(search=user_email)
        else:
            return jsonify({"success": False, "error": "No hay sesión activa"}), 401
            
        return jsonify({"success": True, "tickets": tickets})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets/<identifier>", methods=["GET"])
def get_single_ticket_endpoint(identifier):
    try:
        ticket = database.get_ticket_by_id_or_code(identifier)
        if not ticket:
            return jsonify({"success": False, "error": "Ticket no encontrado"}), 404
        return jsonify({"success": True, "ticket": ticket})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets/<int:ticket_id>", methods=["PATCH"])
def update_ticket_endpoint(ticket_id):
    try:
        data = request.get_json() or {}
        author = data.get("author", "Administrador")
        
        old_ticket = database.get_ticket_by_id_or_code(ticket_id)
        if not old_ticket:
            return jsonify({"success": False, "error": "Ticket no encontrado"}), 404

        updated_ticket = database.update_ticket(ticket_id, data, author=author)
        if not updated_ticket:
            return jsonify({"success": False, "error": "Error al actualizar ticket"}), 404

        # Enviar notificación por correo si cambió de estado o se agregaron notas
        try:
            old_status = old_ticket["status"]
            new_status = updated_ticket["status"]
            old_notes = (old_ticket.get("resolution_notes") or "").strip()
            new_notes = (updated_ticket.get("resolution_notes") or "").strip()
            
            if old_status != new_status or (new_notes and new_notes != old_notes):
                base_url = request.host_url.rstrip('/')
                notifier.notify_ticket_status_updated(updated_ticket, old_status, new_status, base_url)
        except Exception:
            pass
            
        return jsonify({"success": True, "ticket": updated_ticket})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Comentarios & Chat Bidireccional ───────────────────────────────────

@app.route("/api/tickets/<identifier>/comments", methods=["GET"])
def get_ticket_comments_endpoint(identifier):
    try:
        ticket = database.get_ticket_by_id_or_code(identifier)
        if not ticket:
            return jsonify({"success": False, "error": "Ticket no encontrado"}), 404
        mark_read_for = request.args.get("mark_read_for")
        comments = database.get_ticket_comments(ticket["id"], mark_read_for=mark_read_for)
        return jsonify({"success": True, "comments": comments})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/notifications", methods=["GET"])
def get_notifications_endpoint():
    try:
        role = request.args.get("role", "admin")
        user_email = request.args.get("email")
        user_id = session.get("user_id") or request.args.get("user_id")
        
        data = database.get_bell_notifications(role=role, user_email=user_email, user_id=user_id)
        return jsonify({
            "success": True,
            "ticket_groups": data.get("ticket_groups", []),
            "total_count": data.get("total_count", 0),
            "unread_count": data.get("unread_count", 0)
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/notifications/mark-read", methods=["POST"])
def mark_notifications_read_endpoint():
    try:
        data = request.get_json() or {}
        notification_id = data.get("notification_id")
        ticket_id = data.get("ticket_id")
        role = data.get("role", "admin")
        user_email = data.get("email")
        user_id = session.get("user_id") or data.get("user_id")
        
        database.mark_bell_notification_read(
            notification_id=notification_id,
            ticket_id=ticket_id,
            role=role,
            user_email=user_email,
            user_id=user_id
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/notifications/delete", methods=["POST"])
def delete_notification_endpoint():
    try:
        data = request.get_json() or {}
        notification_id = data.get("notification_id")
        ticket_id = data.get("ticket_id")
        role = data.get("role", "admin")
        user_email = data.get("email")
        user_id = session.get("user_id") or data.get("user_id")
        
        database.delete_bell_notification(
            notification_id=notification_id,
            ticket_id=ticket_id,
            role=role,
            user_email=user_email,
            user_id=user_id
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/notifications/clear-all", methods=["POST"])
def clear_all_notifications_endpoint():
    try:
        data = request.get_json() or {}
        role = data.get("role", "admin")
        user_email = data.get("email")
        user_id = session.get("user_id") or data.get("user_id")
        
        database.clear_all_bell_notifications(
            role=role,
            user_email=user_email,
            user_id=user_id
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets/<identifier>/comments", methods=["POST"])
def add_ticket_comment_endpoint(identifier):
    try:
        ticket = database.get_ticket_by_id_or_code(identifier)
        if not ticket:
            return jsonify({"success": False, "error": "Ticket no encontrado"}), 404

        data = request.get_json() or {}
        author_name = data.get("author_name", "Usuario")
        author_email = data.get("author_email", "")
        author_role = data.get("author_role", "usuario")
        message = data.get("message", "")
        image_base64 = data.get("image_base64")
        image_url = data.get("image_url", "")

        if image_base64 and image_base64.startswith("data:image/"):
            try:
                # Extraer formato y datos base64
                header, encoded = image_base64.split(",", 1)
                mime_type = header.split(";")[0].split(":")[1]
                ext = mime_type.split("/")[1] if "/" in mime_type else "png"
                if ext == "jpeg":
                    ext = "jpg"
                
                raw_bytes = base64.b64decode(encoded)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                short_uid = uuid.uuid4().hex[:6]
                orig_name = f"captura_{ts}_{short_uid}.{ext}"
                stored_filename = f"{ticket['code']}_{orig_name}"
                disk_path = os.path.join(app.config["UPLOAD_FOLDER"], stored_filename)
                
                with open(disk_path, "wb") as f:
                    f.write(raw_bytes)
                
                rel_path = os.path.relpath(disk_path, BASE_DIR)
                file_size = len(raw_bytes)
                
                att = database.add_attachment(
                    ticket_id=ticket["id"],
                    original_name=orig_name,
                    stored_name=stored_filename,
                    file_size=file_size,
                    mime_type=mime_type,
                    file_path=rel_path
                )
                image_url = f"/api/attachments/{att['id']}/view"
            except Exception as img_err:
                print(f"Error procesando captura pegada: {img_err}")

        comment = database.add_ticket_comment(
            ticket_id=ticket["id"],
            author_name=author_name,
            author_email=author_email,
            author_role=author_role,
            message=message,
            image_url=image_url
        )

        # Disparar notificación por correo
        try:
            base_url = request.host_url.rstrip('/')
            notifier.notify_ticket_comment(ticket, comment, base_url)
        except Exception:
            pass

        return jsonify({"success": True, "comment": comment}), 201
    except ValueError as ve:
        return jsonify({"success": False, "error": str(ve)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Calificación CSAT (Satisfacción del Cliente) ───────────────────────

@app.route("/api/tickets/<identifier>/rate", methods=["POST"])
def rate_ticket_endpoint(identifier):
    try:
        ticket = database.get_ticket_by_id_or_code(identifier)
        if not ticket:
            return jsonify({"success": False, "error": "Ticket no encontrado"}), 404

        data = request.get_json() or {}
        rating = data.get("rating")
        feedback = data.get("feedback_comment", "")

        updated = database.rate_ticket(ticket["id"], rating, feedback)
        return jsonify({"success": True, "ticket": updated, "message": "¡Gracias por tu calificación!"})
    except ValueError as ve:
        return jsonify({"success": False, "error": str(ve)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Registro de Horas & Time Tracking (Admin) ───────────────────────────

@app.route("/api/tickets/<identifier>/time-logs", methods=["POST"])
def add_time_log_endpoint(identifier):
    try:
        ticket = database.get_ticket_by_id_or_code(identifier)
        if not ticket:
            return jsonify({"success": False, "error": "Ticket no encontrado"}), 404

        data = request.get_json() or {}
        minutes = data.get("minutes")
        description = data.get("description", "")
        author = data.get("author", "Marcelo Ramírez")

        updated = database.add_time_entry(ticket["id"], author=author, minutes=minutes, description=description)
        return jsonify({"success": True, "ticket": updated, "message": "Tiempo registrado exitosamente"})
    except ValueError as ve:
        return jsonify({"success": False, "error": str(ve)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Analíticas Avanzadas (Admin) ───────────────────────────────────────

@app.route("/api/admin/analytics", methods=["GET"])
def admin_analytics_endpoint():
    try:
        analytics = database.get_analytics_data()
        return jsonify({"success": True, "analytics": analytics})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Copia de Seguridad Completa (Backup en 1 Clic) ────────────────────

@app.route("/api/admin/backup", methods=["GET"])
def admin_backup_endpoint():
    try:
        zip_buffer = io.BytesIO()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            db_path = database.DB_PATH
            if os.path.exists(db_path):
                zip_file.write(db_path, arcname="helpdesk.db")

            uploads_dir = app.config.get("UPLOAD_FOLDER", os.path.join(BASE_DIR, "uploads"))
            if os.path.exists(uploads_dir):
                for root, dirs, files in os.walk(uploads_dir):
                    for file in files:
                        full_path = os.path.join(root, file)
                        rel_path = os.path.relpath(full_path, BASE_DIR)
                        zip_file.write(full_path, arcname=rel_path)

        zip_buffer.seek(0)
        filename = f"AquaShield_Backup_{timestamp}.zip"
        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Descargas y Visualización de Adjuntos ──────────────────────────────

@app.route("/api/attachments/<int:attachment_id>/view", methods=["GET"])
def view_attachment_endpoint(attachment_id):
    try:
        att = database.get_attachment_by_id(attachment_id)
        if not att:
            abort(404, description="Archivo adjunto no encontrado")
            
        file_full_path = os.path.join(BASE_DIR, att["file_path"])
        if not os.path.exists(file_full_path):
            abort(404, description="El archivo físico no se encuentra en el servidor")
            
        return send_file(
            file_full_path,
            as_attachment=False,
            mimetype=att["mime_type"] or "image/png"
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets/<int:ticket_id>/attachments/<int:attachment_id>/download", methods=["GET"])
def download_single_attachment(ticket_id, attachment_id):
    try:
        att = database.get_attachment_by_id(attachment_id)
        if not att or att["ticket_id"] != ticket_id:
            abort(404, description="Archivo adjunto no encontrado")
            
        file_full_path = os.path.join(BASE_DIR, att["file_path"])
        if not os.path.exists(file_full_path):
            abort(404, description="El archivo físico no se encuentra en el servidor")
            
        return send_file(
            file_full_path,
            as_attachment=True,
            download_name=att["original_name"],
            mimetype=att["mime_type"] or "application/octet-stream"
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets/<identifier>/attachments", methods=["POST"])
def upload_ticket_attachment_endpoint(identifier):
    try:
        ticket = database.get_ticket_by_id_or_code(identifier)
        if not ticket:
            return jsonify({"success": False, "error": "Ticket no encontrado"}), 404

        uploaded_files = request.files.getlist("attachments") or request.files.getlist("files[]") or request.files.getlist("file")
        if not uploaded_files:
            return jsonify({"success": False, "error": "No se recibieron archivos"}), 400

        ticket_upload_dir = os.path.join(app.config["UPLOAD_FOLDER"], ticket["code"])
        os.makedirs(ticket_upload_dir, exist_ok=True)

        added_attachments = []
        for file in uploaded_files:
            if file and file.filename and file.filename.strip():
                orig_name = os.path.basename(file.filename)
                safe_name = secure_filename(orig_name) or "adjunto"
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                stored_name = f"{timestamp}_{safe_name}"
                disk_path = os.path.join(ticket_upload_dir, stored_name)
                file.save(disk_path)

                file_size = os.path.getsize(disk_path)
                mime_type = mimetypes.guess_type(disk_path)[0] or "application/octet-stream"
                rel_path = os.path.relpath(disk_path, BASE_DIR)

                att = database.add_attachment(
                    ticket_id=ticket["id"],
                    original_name=orig_name,
                    stored_name=stored_name,
                    file_size=file_size,
                    mime_type=mime_type,
                    file_path=rel_path
                )
                added_attachments.append(att)

        return jsonify({"success": True, "attachments": added_attachments}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/tickets/<int:ticket_id>/download-zip", methods=["GET"])
def download_ticket_zip(ticket_id):
    try:
        ticket = database.get_ticket_by_id_or_code(ticket_id)
        if not ticket:
            abort(404, description="Ticket no encontrado")
            
        attachments = ticket.get("attachments", [])
        if not attachments:
            return jsonify({"success": False, "error": "Este ticket no tiene archivos adjuntos"}), 400
            
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            summary_txt = f"""=====================================================
AQUASHIELD · MESA DE AYUDA - AQUACHILE
DETALLE DE REQUERIMIENTO / TICKET
=====================================================
Código Ticket : {ticket['code']}
Fecha Ingreso : {ticket['created_at']}
Solicitante   : {ticket['requester_name']} <{ticket['requester_email']}>
Teléfono/Wsp  : {ticket.get('requester_phone', 'N/A')}
Módulo        : {ticket['module_name']}
Tipo Solicitud: {ticket['type'].upper()}
Prioridad     : {ticket['priority'].upper()}
Estado Actual : {ticket['status'].upper()}

ASUNTO:
{ticket['title']}

DESCRIPCIÓN DETALLADA:
{ticket['description']}

NOTAS DE RESOLUCIÓN:
{ticket.get('resolution_notes', 'Sin notas registradas')}
=====================================================
Total Adjuntos Incluidos: {len(attachments)}
"""
            zip_file.writestr(f"{ticket['code']}_resumen.txt", summary_txt.encode("utf-8"))
            
            for att in attachments:
                file_full_path = os.path.join(BASE_DIR, att["file_path"])
                if os.path.exists(file_full_path):
                    zip_file.write(file_full_path, arcname=f"adjuntos/{att['original_name']}")
                    
        zip_buffer.seek(0)
        zip_filename = f"{ticket['code']}_adjuntos.zip"
        
        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name=zip_filename
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Métricas & Exportación ─────────────────────────────────────────────

@app.route("/api/stats", methods=["GET"])
def get_stats_endpoint():
    try:
        stats = database.get_dashboard_stats()
        return jsonify({"success": True, "stats": stats})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/export/excel", methods=["GET"])
def export_excel_endpoint():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        status = request.args.get("status")
        module_id = request.args.get("module_id")
        ticket_type = request.args.get("type")
        priority = request.args.get("priority")
        search = request.args.get("search")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        tag = request.args.get("tag")

        tickets = database.get_tickets(
            status=status,
            module_id=module_id,
            ticket_type=ticket_type,
            priority=priority,
            search=search,
            date_from=date_from,
            date_to=date_to,
            tag=tag,
            limit=10000
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets AquaShield"
        
        header_fill = PatternFill(start_color="445563", end_color="445563", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        headers = [
            "Código", "Fecha Creación", "Solicitante", "Email", "Teléfono",
            "Módulo", "Tipo", "Prioridad", "Estado", "Etiquetas", "Asunto", "Descripción", 
            "Adjuntos", "Mensajes Chat", "Calificación CSAT", "Comentario Usuario", "Notas Resolución"
        ]
        
        ws.append(headers)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for t in tickets:
            ws.append([
                t["code"],
                t["created_at"],
                t["requester_name"],
                t["requester_email"],
                t.get("requester_phone", ""),
                t["module_name"],
                t["type"].upper(),
                t["priority"].upper(),
                t["status"].upper(),
                t.get("tags", ""),
                t["title"],
                t["description"],
                t.get("attachments_count", 0),
                t.get("comments_count", 0),
                f"{t.get('rating')} Estrellas" if t.get("rating") else "Sin calificar",
                t.get("feedback_comment", ""),
                t.get("resolution_notes", "")
            ])
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
            
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"AquaShield_Mesa_Ayuda_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            excel_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/admin/report-executive", methods=["GET"])
def executive_report_page():
    try:
        analytics = database.get_analytics_data()
        all_tickets = database.get_tickets(limit=1000)
        resolved_tickets = [t for t in all_tickets if t["status"] == "resuelto"]
        
        return render_template(
            "executive_report.html",
            analytics=analytics,
            tickets=all_tickets,
            resolved_tickets=resolved_tickets,
            generated_at=datetime.now().strftime("%d/%m/%Y %H:%M")
        )
    except Exception as e:
        return f"Error generando reporte: {str(e)}", 500

@app.route("/manifest.json", methods=["GET"])
def pwa_manifest():
    return send_from_directory("static", "manifest.json", mimetype="application/manifest+json")

@app.route("/sw.js", methods=["GET"])
def pwa_service_worker():
    return send_from_directory("static/js", "sw.js", mimetype="application/javascript")

# ── ENDPOINTS DE LA BANDEJA DE ESPERA (GITHUB QUEUE) ────────────────────────
@app.route("/api/queue/status", methods=["GET"])
def api_queue_status():
    try:
        import sync_github_queue
        issues = sync_github_queue.fetch_pending_github_issues()
        return jsonify({"count": len(issues), "issues": issues})
    except Exception as e:
        return jsonify({"count": 0, "error": str(e)})

@app.route("/api/queue/sync", methods=["POST"])
def api_queue_sync():
    try:
        import sync_github_queue
        author = request.json.get("author", "Administrador") if request.is_json else "Administrador"
        result = sync_github_queue.sync_github_queue_to_db(author=author)
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def get_local_ip():
    try:
        import socket
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

if __name__ == "__main__":
    import socket
    port = int(os.environ.get("PORT", 5050))
    local_ip = get_local_ip()
    hostname = socket.gethostname()
    fqdn = socket.getfqdn()
    
    print("=" * 70)
    print("       AQUASHIELD - MESA DE AYUDA (AquaChile)")
    print("=" * 70)
    print(f"[*] 1. ENLACE CORPORATIVO RECOMENDADO (Red / VPN):")
    print(f"       http://{fqdn}:{port}")
    print(f"\n[*] 2. ENLACE POR IP DIRECTA (Si el DNS no resuelve el nombre):")
    print(f"       http://{local_ip}:{port}")
    print(f"\n[*] 3. ENLACE POR NOMBRE CORTO DE PC:")
    print(f"       http://{hostname}:{port}")
    print(f"\n[*] TU ACCESO LOCAL (En tu PC):")
    print(f"       http://localhost:{port}  (Panel Admin: http://localhost:{port}/admin)")
    print("=" * 70)
    print("Los usuarios normales NO necesitan instalar nada. Solo abren el enlace.")
    print("=" * 70 + "\n")
    
    app.run(host="0.0.0.0", port=port, debug=False)
